import pandas as pd
import psycopg2
from openpyxl import Workbook
from config import load_config


config = load_config()
names = []
comments = []
try:
    with psycopg2.connect(**config) as conn:
        cur = conn.cursor()
        cur.execute('select name, comment from snap_v1.perfume')
        row = cur.fetchone()
        while row is not None:
            names.append(row[0])
            comments.append(row[1])
            row = cur.fetchone()

        conn.commit()
        cur.close()
except psycopg2.DatabaseError as error:
    if conn:
        conn.rollback()
    print(error)
finally:
    if conn:
        conn.close()

workbook = Workbook()
sheet = workbook.active

for row, (name, comment) in enumerate(zip(names, comments), start=1):
    sheet.cell(row, 1, name)       # Колонка A
    sheet.cell(row, 2, comment)    # Колонка B

# Сохраняем созданную книгу Excel
workbook.save('data.xlsx')